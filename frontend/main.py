import subprocess
import sys
import pkg_resources

def install_libraries():
    required_libraries = ["google-generativeai", "Flask", "flask-cors","openpyxl"]

    installed_libraries = {package.project_name.lower(): package.version for package in pkg_resources.working_set}
    
    missing_libraries = [lib for lib in required_libraries if lib.lower() not in installed_libraries]

    if missing_libraries:
        subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing_libraries)
    else:
        print("")

# Call the function to install libraries if necessary
install_libraries()
from flask import Flask, request, jsonify
from flask_cors import CORS
import google.generativeai as palm
import openpyxl
from datetime import datetime

app = Flask(__name__)
CORS(app) 

@app.route('/generate_response', methods=['POST'])
def generate_response():
    data = request.get_json()
    prompt = data.get('prompt', '')

    apikey = "AIzaSyCVrjd8JXl1CDUlsuKxqBjBojieVcPaRUk"
    palm.configure(api_key=apikey)
    model_id = "models/text-bison-001"

    compl = palm.generate_text(
        model=model_id,
        prompt=prompt,
        temperature=0.99,
        max_output_tokens=800,
        candidate_count=1
    )

    outputs = [output["output"] for output in compl.candidates]
    response = outputs[0] if outputs else "No response"

    # Save to Excel file
    log_to_excel(prompt, response)

    return jsonify({'response': response})

@app.route('/contact_us', methods=['POST'])
def contact_us():
    data = request.get_json()
    firstName= data.get('firstName', '')
    lastName= data.get('lastName', '')
    email= data.get('email', '')
    mobile= data.get('mobile', '')
    messag= data.get('message', '')

    # Save to Excel file
    log_to_excel_contact(firstName,lastName,email,mobile,messag)



def log_to_excel(prompt, response):
    # Get current date and time
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Load or create the Excel file
    try:
        workbook = openpyxl.load_workbook("ws.xlsx")
        print("Workbook found")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.active.append(["Prompt", "Response", "Date"])

    # Select the active sheet (assume it's the first sheet)
    sheet = workbook.active
    print(sheet)

    # Add headers if the sheet is empty
    
    # Append new data
    next_row = sheet.max_row + 1
    print(next_row)
    sheet.cell(row=next_row, column=1, value=prompt)
    sheet.cell(row=next_row, column=2, value=response)
    sheet.cell(row=next_row, column=3, value=current_time)

    # Save the workbook
    workbook.save("ws.xlsx")

def log_to_excel_contact(firstName,lastName,email,mobile,messag):
    # Get current date and time
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Load or create the Excel file
    try:
        workbook = openpyxl.load_workbook("contact_us.xlsx")
        print("Workbook found")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.active.append(["FName", "LName", "Email","Mobile","Message","Date"])

    # Select the active sheet (assume it's the first sheet)
    sheet = workbook.active
    print(sheet)

    # Add headers if the sheet is empty
    
    # Append new data
    next_row = sheet.max_row + 1
    print(next_row)
    sheet.cell(row=next_row, column=1, value=firstName)
    sheet.cell(row=next_row, column=2, value=lastName)
    sheet.cell(row=next_row, column=3, value=email)
    sheet.cell(row=next_row, column=4, value=mobile)
    sheet.cell(row=next_row, column=5, value=messag)
    sheet.cell(row=next_row, column=6, value=current_time)

    # Save the workbook
    workbook.save("contact_us.xlsx")

if __name__ == '__main__':
    app.run(debug=True)